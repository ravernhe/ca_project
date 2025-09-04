# -*- coding: utf-8 -*-

from __future__ import annotations
import re, unicodedata, csv
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from .settings import DATE_COLUMNS

# ---------- normalization ----------
def norm(s: str) -> str:
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def build_col_index(columns) -> dict:
    return {norm(c): c for c in columns}

def find_col(idx: dict, columns, *, any_of=(), all_tokens=(), regex: Optional[str]=None):
    for name in any_of:
        key = norm(name)
        if key in idx:
            return idx[key]
    if all_tokens:
        toks = [norm(t) for t in all_tokens]
        for c in columns:
            nc = norm(c)
            if all(t in nc for t in toks):
                return c
    if regex:
        cre = re.compile(regex)
        for c in columns:
            if cre.search(norm(c)):
                return c
    return None

# ---------- IO ----------
def read_any(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path)
    if path.suffix.lower() == ".csv":
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
        except Exception:
            try:
                return pd.read_csv(path, sep=";", engine="python", encoding="utf-8-sig")
            except Exception:
                return pd.read_csv(path, sep=",", engine="python", encoding="utf-8-sig")
    raise ValueError(f"Unsupported file type: {path.suffix}")

def read_csv_loose(path: Path) -> pd.DataFrame:
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            sample = f.read(4096)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        delim = dialect.delimiter
    except Exception:
        delim = None

    try:
        return pd.read_csv(path, header=None, sep=None, engine="python",
                           quoting=csv.QUOTE_MINIMAL, quotechar='"', escapechar="\\",
                           on_bad_lines="skip", encoding="utf-8-sig")
    except Exception:
        pass

    if delim:
        try:
            return pd.read_csv(path, header=None, sep=delim, engine="python",
                               quoting=csv.QUOTE_MINIMAL, quotechar='"', escapechar="\\",
                               on_bad_lines="skip", encoding="utf-8-sig")
        except Exception:
            pass

    for d in [";", ",", "\t", "|"]:
        try:
            return pd.read_csv(path, header=None, sep=d, engine="python",
                               quoting=csv.QUOTE_MINIMAL, quotechar='"', escapechar="\\",
                               on_bad_lines="skip", encoding="utf-8-sig")
        except Exception:
            continue

    try:
        return pd.read_table(path, header=None, sep=r"[;,\t|]", engine="python",
                             on_bad_lines="skip", encoding="utf-8-sig")
    except Exception as e:
        raise ValueError(f"CSV illisible: {path.name} ({e})")

def export_excel(df: pd.DataFrame, out_path: Path, meta: dict):
    ws_name = "consolidation"

    # --- Colonnes à cibler ---
    date_cols = [
        "Date d'éxécution (produit sans session)",
        "Date de début",
        "Date de fin",
        "Date prévisionnelle de début de projet",
        "Date de fin de projet",
        "Date de début repère",
        "Date de fin repère",
    ]
    price_cols = [
        "Prix Intra 1 standard (converti)",
        "Prix total",
        "CA session",
        "Facturation (convertie) N-2",
        "Facturation (convertie) N-1",
        "Facturation (convertie) N",
        "Facturation Y-2",
        "Facturation Y-1",
        "Facturation Y",
        "Facturation totale",
        "CA attendu",
        "CA avancement",
        "CA YTD",
        "CA EOY (backlog)",
        "FAE",
        "PCA",
        "Prix de vente (converti)",
        "Prix total (converti)",
    ]
    percent_cols = ["Taux d'avancement global", "Avancement EOY"]

    # --- Préparation du DataFrame ---
    df_out = df.copy()

    # convertir les colonnes de dates en datetime64
    for col in date_cols:
        if col in df_out.columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce")

    # remplacer NaN uniquement sur les colonnes non-dates
    non_date_cols = [c for c in df_out.columns if c not in date_cols]
    df_out[non_date_cols] = df_out[non_date_cols].replace(
        {pd.NA: "", "nan": "", "<NA>": "", None: ""}
    )

    # --- Écriture des feuilles ---
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df_out.to_excel(xw, sheet_name=ws_name, index=False)
        resume = (
            df_out.groupby("Origine rapport", dropna=False)
            .size()
            .rename("nb_lignes")
            .reset_index()
        )
        resume.to_excel(xw, sheet_name="resume", index=False)
        pd.DataFrame([meta]).to_excel(xw, sheet_name="parametres", index=False)

    # --- Mise en forme Excel ---
    wb = load_workbook(out_path)

    # auto-largeur en fonction de l'en-tête
    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value is not None:
                col_letter = cell.column_letter
                ws.column_dimensions[col_letter].width = len(str(cell.value)) + 2

    ws = wb[ws_name]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    # Dates → DD/MM/YY
    for name in date_cols:
        if name in headers:
            ci = headers[name]
            for col in ws.iter_cols(
                min_col=ci, max_col=ci, min_row=2, values_only=False
            ):
                for cell in col:
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = "dd/mm/yy"

    # Prix → #
    for name in price_cols:
        if name in headers:
            ci = headers[name]
            for col in ws.iter_cols(
                min_col=ci, max_col=ci, min_row=2, values_only=False
            ):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#"

    # Pourcentages → 0%
    for name in percent_cols:
        if name in headers:
            ci = headers[name]
            for col in ws.iter_cols(
                min_col=ci, max_col=ci, min_row=2, values_only=False
            ):
                for cell in col:
                    cell.number_format = "0%"

    wb.save(out_path)



# ---------- transformations ----------
def _clean_str_series(s: pd.Series) -> pd.Series:
    if not isinstance(s, pd.Series):
        return pd.Series([], dtype="object")
    return (
        s.astype(str)
         .str.replace("\u00A0", " ", regex=False)   # espace insécable
         .str.strip()
         .replace({"": pd.NA})
    )

def _to_dt_fr(s: pd.Series) -> pd.Series:
    """
    Parse FR pendant le parsing :
      - JJ/MM/AAAA
      - JJ/MM/AA
    Tolère aussi '-' et '.' comme séparateurs et nombres Excel.
    Ne tente PAS le format US (MDY).
    """
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce")

    s_clean = _clean_str_series(s)
    s_norm  = s_clean.str.replace(r"[-\.]", "/", regex=True)

    out = pd.to_datetime(s_norm, format="%d/%m/%Y", errors="coerce")  # JJ/MM/AAAA
    need = out.isna()
    if need.any():
        out.loc[need] = pd.to_datetime(s_norm[need], format="%d/%m/%y", errors="coerce")  # JJ/MM/AA

    need = out.isna()
    if need.any():
        # fallback FR (dayfirst)
        out.loc[need] = pd.to_datetime(s_norm[need], errors="coerce", dayfirst=True, yearfirst=False)

    # nombres Excel éventuels
    need = out.isna()
    if need.any():
        num_mask = pd.to_numeric(s_norm[need], errors="coerce").notna()
        idx = s_norm[need][num_mask].index
        if len(idx):
            nums = pd.to_numeric(s_norm.loc[idx], errors="coerce")
            out.loc[idx] = pd.to_datetime(nums, unit="D", origin="1899-12-30", errors="coerce")

    return out

# ---- nouvelle version ----

def format_date_columns(df: pd.DataFrame, style: str = "datetime_fr") -> pd.DataFrame:
    """
    - 'datetime_fr' (par défaut) : parse les colonnes DATE_COLUMNS en datetime (FR), sans formater en texte.
    - 'dd/mm/yy' : parse FR puis rend sous forme texte 'dd/mm/yy' (utile seulement si tu veux voir du texte).
    - 'iso' : parse FR puis rend 'YYYY-MM-DD' (texte).
    """
    df = df.copy()

    for col in DATE_COLUMNS:
        if col in df.columns:
            s = _to_dt_fr(df[col])
            if style == "datetime_fr":
                df[col] = s
            elif style == "dd/mm/yy":
                df[col] = s.dt.strftime("%d/%m/%y").fillna("")
            elif style == "iso":
                df[col] = s.dt.strftime("%Y-%m-%d").fillna("")
            else:
                # fallback : garder datetime
                df[col] = s
    return df

def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df